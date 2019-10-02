#import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

pixels = pd.read_csv('train_digit_recognizer.csv').as_matrix()
pixels_test = pd.read_csv('test_digit_recognizer.csv').as_matrix()


X_train = pixels[:,1:]
Y_train = pixels[:,0]

X_test = pixels_test[:,:]
#Y_test = pixels_test[:,0]

#from sklearn.tree import DecisionTreeClassifier
#clf = DecisionTreeClassifier()
#clf.fit(X_train,Y_train)

from sklearn.ensemble import RandomForestClassifier
clf2 = RandomForestClassifier(n_estimators=1500,random_state=0)
clf2.fit(X_train,Y_train)

#Y_pred = clf.predict(X_test)
Y_pred2 = clf2.predict(X_test)

Y_pred2_list = Y_pred2.tolist()
##type(list(Y_pred2[0]))
import openpyxl as xl

wb = xl.Workbook()
sheet = wb.active
#sheet['A1'].value = 1
from openpyxl.utils import get_column_letter
#get_column_letter(1)

j=1
for i in Y_pred2_list:
    sheet['A' + str(j)] = j
    sheet['B' + str(j)] = i
    j=j+1

wb.save('preds2.xlsx')

#import csv
#
#with open('preds.csv','w') as csvw:
#    csvwriter = csv.writer(csvw)
#    for i in Y_pred2_list:
#        csvwriter.writerow([i])

#csvw = open('preds.csv','w')
#csvwriter = csv.writer(csvw)
#
##for i in list(Y_pred2_list):
##    csvwriter.writerow([i])
#csvwriter.writerows(Y_pred2_list)
#    
#csvw.close()

#print(clf.predict([X_train[7]]))
#d = X_train[7]
#d.shape=(28,28)
#plt.imshow(255-d,cmap= 'gray')
#plt.show()
#
#from sklearn.metrics import accuracy_score
##print("Acc",accuracy_score(Y_test,Y_pred))
#print("Accuracy : ",accuracy_score(Y_test,Y_pred2))
